import re
import math
import json
import logging
import argparse
import threading
import traceback
import unicodedata
from pathlib import Path
from typing import Dict, List, Tuple, Optional, Any
from functools import lru_cache

import pandas as pd
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
from openpyxl.styles import PatternFill

try:
    from rapidfuzz import fuzz, process
except ImportError:
    from difflib import SequenceMatcher

    class _Fuzz:
        @staticmethod
        def ratio(a, b):
            return int(100 * SequenceMatcher(None, a, b).ratio())

        @staticmethod
        def partial_ratio(a, b):
            if not a or not b:
                return 0
            a, b = (a, b) if len(a) <= len(b) else (b, a)
            best = 0
            for i in range(max(1, len(b) - len(a) + 1)):
                best = max(best, SequenceMatcher(None, a, b[i:i+len(a)]).ratio())
            return int(100 * best)

        @staticmethod
        def token_set_ratio(a, b):
            sa = set(a.split())
            sb = set(b.split())
            inter = ' '.join(sorted(sa & sb))
            left = ' '.join(sorted(sa))
            right = ' '.join(sorted(sb))
            return max(
                _Fuzz.ratio(inter, left),
                _Fuzz.ratio(inter, right),
                _Fuzz.ratio(left, right)
            )

    fuzz = _Fuzz()
    process = None

# ------------------------------
# Logging setup
# ------------------------------
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

# ------------------------------
# Default configuration (can be overridden by config.json)
# ------------------------------
DEFAULT_CONFIG = {
    "master_weights": {
        "company_name": 25,
        "group_name": 10,
        "facility_name": 20,
        "facility_type": 10,
        "full_address": 20,
        "city_name": 5,
        "state_name": 5,
        "country_name": 5,
    },
    "confidence_buckets": [
        [90, "High confidence"],
        [75, "Medium confidence"],
        [60, "Low confidence / Review recommended"],
        [0, "No Match"]
    ],
    "penalty_values": {
        "facility_type_conflict": 10,
        "company_mismatch_weak_address": 18,
        "feature_entity_no_support": 10,
        "explicit_company_mismatch": 12,
        "address_stronger_than_entity": 8,
        "company_match_no_site_evidence": 14,
        "country_conflict": 18,
        "state_conflict": 8,
        "address_geography_conflict": 10,
        "code_mismatch": 35,
        "candidate_missing_code": 8,
        "candidate_missing_code_weak": 20,
        "code_and_address_aligned_bonus": -8
    },
    "score_thresholds": {
        "shortlist_min_score": 35,
        "high_confidence": 90,
        "medium_confidence": 75,
        "low_confidence": 60,
        "needs_review_threshold": 75
    },
    "address_similarity": {
        "no_numeric_penalty": 0.72,
        "numeric_mismatch_penalty": 0.78,
        "first_num_mismatch_penalty": 0.72,
        "numeric_score_weight": 0.1,
        "overlap_weight": 0.2,
        "base_weight": 0.7
    },
    "shortlist_overlap_weights": {
        "feature_token": 1.2,
        "address_token": 2.2,
        "facility_token": 3.0,
        "manufacturer_token": 3.0,
        "supplier_token": 2.5,
        "company_hint_token": 3.2,
        "entity_match_bonus": 12,
        "manufacturer_match_bonus": 12,
        "supplier_match_bonus": 12,
        "facility_hint_match_bonus": 12,
        "address_hint_match_bonus": 12,
        "code_match_bonus": 15
    },
    "facility_type_groups": {
        "fabrication": {"fabrication", "fab", "wafer fab", "foundry", "front end"},
        "factory": {"factory", "plant", "manufacturing"},
        "ic assembly": {"assembly", "ic assembly", "assembly and test", "backend", "back end", "osat", "test"},
        "office": {"office", "sales office"},
        "headquarters": {"headquarters", "hq"},
        "design": {"design center", "r and d center", "r&d center", "r d center", "design"},
        "warehouse": {"warehouse"}
    },
    "company_suffixes": [
        "co", "co.", "ltd", "ltd.", "limited", "inc", "inc.", "corporation", "corp", "corp.",
        "llc", "gmbh", "ag", "plc", "s.a.", "sa", "pte", "pte.", "bv", "b.v.", "lp", "holdings",
        "holding", "technology", "technologies", "electronics", "electronic", "semiconductor",
        "semiconductors", "industry", "industrial", "devices", "device", "company", "public", "private"
    ],
    "generic_tokens": [
        "the", "and", "road", "street", "avenue", "boulevard", "drive", "lane", "highway", "suite",
        "number", "building", "district", "province", "city", "county", "industrial", "park", "zone",
        "factory", "fabrication", "assembly", "office", "campus", "site", "group", "technology",
        "technologies", "electronics", "electronic", "semiconductor", "semiconductors", "company",
        "corporation", "limited", "incorporated", "plant", "manufacturer"
    ],
    "country_synonyms": {
        "china": ["china", "pr china", "mainland china", "cn"],
        "taiwan": ["taiwan", "taiwan roc", "roc", "r o c"],
        "south korea": ["south korea", "korea", "republic of korea", "korea republic of"],
        "japan": ["japan"],
        "thailand": ["thailand", "th"],
        "israel": ["israel"],
        "malaysia": ["malaysia"],
        "united states": ["united states", "usa", "us", "u s a"],
        "united kingdom": ["uk", "united kingdom", "u k"]
    },
    "acronym_aliases": {
        "tsmc": "taiwan semiconductor manufacturing company",
        "ase": "ase technology holding",
        "asek": "ase technology holding",
        "onsemi": "onsemi",
        "ti": "texas instruments",
        "kla": "kla corporation",
        "utac": "utac",
        "hana": "hana microelectronics",
        "tower": "tower semiconductor"
    },
    "abbreviation_match_threshold": 85
}

# Load config from file if present
CONFIG_PATH = Path(__file__).parent / "location_mapping_config.json"
if CONFIG_PATH.exists():
    with open(CONFIG_PATH, "r") as f:
        user_config = json.load(f)
        DEFAULT_CONFIG.update(user_config)

MASTER_WEIGHTS = DEFAULT_CONFIG["master_weights"]
CONFIDENCE_BUCKETS = DEFAULT_CONFIG["confidence_buckets"]
PENALTIES = DEFAULT_CONFIG["penalty_values"]
SCORE_THRESHOLDS = DEFAULT_CONFIG["score_thresholds"]
ADDR_SIM = DEFAULT_CONFIG["address_similarity"]
SHORTLIST_W = DEFAULT_CONFIG["shortlist_overlap_weights"]
FACILITY_TYPE_GROUPS = DEFAULT_CONFIG["facility_type_groups"]
COMPANY_SUFFIXES = set(DEFAULT_CONFIG["company_suffixes"])
GENERIC_TOKENS = set(DEFAULT_CONFIG["generic_tokens"])
COUNTRY_SYNONYMS = DEFAULT_CONFIG["country_synonyms"]
ACRONYM_ALIASES = DEFAULT_CONFIG["acronym_aliases"]
ABBREVIATION_MATCH_THRESHOLD = DEFAULT_CONFIG["abbreviation_match_threshold"]

# ------------------------------
# Constants
# ------------------------------
COLUMN_ALIASES = {
    'location_id': ['location_id', 'location id', 'loc_id', 'locationid', 'companylocationid', 'company location id'],
    'company_name': ['company_name', 'company name', 'company'],
    'group_name': ['groupname', 'group_name', 'group name', 'parent company', 'group'],
    'facility_name': ['facility_name', 'facility name', 'site name', 'plant name', 'location name'],
    'facility_type': ['facility_type', 'facility type', 'facilitytype', 'site level', 'sitelevel'],
    'country_name': ['countryname', 'country_name', 'country name', 'country'],
    'full_address': ['full_address', 'full address', 'address', 'site address', 'location address'],
    'state_name': ['state_name', 'state name', 'province', 'state'],
    'city_name': ['city_name', 'city name', 'city'],
    'feature_value': ['featurevalue', 'feature value', 'feature', 'value'],
    'input_facility_type': ['facilitytype', 'facility type', 'facility_type'],
    'input_row_id': ['input_row_id', 'input row id', 'row id', 'id'],
    'manufacturer': ['manufacturer'],
    'contacted_supplier': ['contactedsupplier', 'contacted supplier', 'supplier'],
    'input_country': ['country', 'countryname'],
    'input_state': ['stateprovince', 'state province', 'state', 'province'],
    'input_city': ['city', 'city_name', 'city name'],
}

ADDRESS_ABBREVIATIONS = {
    'st': 'street', 'rd': 'road', 'ave': 'avenue', 'av': 'avenue', 'blvd': 'boulevard',
    'dr': 'drive', 'ln': 'lane', 'hwy': 'highway', 'pkwy': 'parkway', 'ste': 'suite',
    'no': 'number', 'nr': 'number', 'bldg': 'building', 'ctr': 'center', 'ct': 'court'
}

COUNTRY_ALIASES = {k: v for k, v in COUNTRY_SYNONYMS.items() for v in v}
COUNTRY_ALIASES.update({v: k for k, vals in COUNTRY_SYNONYMS.items() for v in vals})

STATE_ALIASES = {
    'ca': 'california', 'tx': 'texas', 'ny': 'new york', 'wa': 'washington', 'il': 'illinois',
    'ma': 'massachusetts', 'az': 'arizona', 'or': 'oregon', 'nm': 'new mexico',
    'mi': 'michigan', 'oh': 'ohio', 'pa': 'pennsylvania', 'nj': 'new jersey'
}

FACILITY_VARIANTS = {
    'hq': 'headquarters', 'plant': 'factory', 'fab': 'fabrication', 'site': 'site',
    'campus': 'campus', 'office': 'office', 'warehouse': 'warehouse', 'factory': 'factory',
    'building': 'building', 'assembly': 'assembly'
}

NON_SITE_ORDINAL_CODES = {'1st', '2nd', '3rd', '4th', '5th', '6th'}

# ------------------------------
# Helper functions
# ------------------------------
def clean_header(s: str) -> str:
    s = str(s).strip().lower()
    s = re.sub(r'[^a-z0-9]+', ' ', s)
    return re.sub(r'\s+', ' ', s).strip()

def strip_accents(text: str) -> str:
    return ''.join(
        c for c in unicodedata.normalize('NFKD', text)
        if not unicodedata.combining(c)
    )

def normalize_text(value) -> str:
    if pd.isna(value):
        return ''
    text = str(value)
    text = strip_accents(text).lower().strip()
    text = text.replace('&', ' and ')
    text = re.sub(r'[_|/]+', ' ', text)
    text = re.sub(r'[,:;#()\[\]{}]', ' ', text)
    text = text.replace('-', ' ')
    text = re.sub(r'\bno\.\b', ' number ', text)
    text = re.sub(r'\s+', ' ', text).strip()

    tokens = []
    for token in text.split():
        token = COUNTRY_ALIASES.get(token, token)
        token = STATE_ALIASES.get(token, token)
        token = ADDRESS_ABBREVIATIONS.get(token, token)
        token = FACILITY_VARIANTS.get(token, token)
        tokens.append(token)
    text = ' '.join(tokens)
    text = re.sub(r'\s+', ' ', text).strip()
    return text

def normalize_company(value) -> str:
    text = normalize_text(value)
    tokens = [ACRONYM_ALIASES.get(t, t) for t in text.split()]
    merged = ' '.join(tokens)
    tokens = [t for t in merged.split() if t not in COMPANY_SUFFIXES]
    return ' '.join(tokens).strip()

def extract_acronyms(value: str) -> set:
    raw = safe_str(value)
    acronyms = set()
    for token in re.findall(r'\b[A-Z]{2,6}\b', raw):
        acronyms.add(token.lower())
    norm = normalize_text(raw)
    for token in norm.split():
        if token in ACRONYM_ALIASES:
            acronyms.add(token)
    if raw:
        initials = ''.join(ch.lower() for ch in re.findall(r'\b[A-Za-z]', raw))
        if 2 <= len(initials) <= 6:
            acronyms.add(initials)
    return acronyms

def token_set(text: str) -> set:
    return {t for t in normalize_text(text).split() if t}

def informative_tokens(text: str) -> set:
    return {t for t in token_set(text) if len(t) >= 3 and t not in GENERIC_TOKENS and not t.isdigit()}

def safe_str(v) -> str:
    return '' if pd.isna(v) else str(v).strip()

def extract_code_tokens(text: str) -> set:
    raw = safe_str(text)
    norm = normalize_text(text)
    codes = set()
    raw_lower = strip_accents(raw).lower()
    prefix_aliases = {
        'fab': 'fab',
        'fabrication': 'fab',
        'utl': 'utl',
        'site': 'site',
        'building': 'building',
        'bldg': 'building',
        'line': 'line',
        'plant': 'plant',
        'factory': 'factory',
        't': 't',
        'a': 'a',
    }
    for token in re.findall(r'\b[a-z]{1,8}\d{1,4}[a-z]?\b|\b\d{1,3}[a-z]{1,5}\b', norm):
        if token not in NON_SITE_ORDINAL_CODES:
            codes.add(token)
    for source in [raw_lower, norm]:
        for prefix, number in re.findall(r'\b(fab|fabrication|utl|site|building|bldg|line|plant|factory|t|a)\s*[- ]?(\d{1,4}[a-z]?)\b', source):
            token = f"{prefix_aliases.get(prefix, prefix)}{number}"
            if token not in NON_SITE_ORDINAL_CODES:
                codes.add(token)
        for number, suffix in re.findall(r'\b(\d{1,4}[a-z]?)\s*[- ]?(fab|fabrication|utl|site|building|bldg|line|plant|factory)\b', source):
            token = f"{prefix_aliases.get(suffix, suffix)}{number}"
            if token not in NON_SITE_ORDINAL_CODES:
                codes.add(token)
    for token in re.findall(r'\b[a-z]{2,8}\b', norm):
        if token in {'chungli', 'millennium'}:
            codes.add(token)
    return codes

def similarity(a: str, b: str) -> float:
    a = normalize_text(a)
    b = normalize_text(b)
    if not a or not b:
        return 0.0
    return max(
        fuzz.ratio(a, b),
        fuzz.partial_ratio(a, b),
        fuzz.token_set_ratio(a, b),
    ) / 100.0

def exact_or_contains(a: str, b: str) -> float:
    a = normalize_text(a)
    b = normalize_text(b)
    if not a or not b:
        return 0.0
    if a == b:
        return 1.0
    if a in b or b in a:
        return 0.92
    return 0.0

def company_similarity(a: str, b: str) -> float:
    a1 = normalize_company(a)
    b1 = normalize_company(b)
    if not a1 or not b1:
        return 0.0
    acr_a = {ACRONYM_ALIASES.get(x, x) for x in extract_acronyms(a)}
    acr_b = {ACRONYM_ALIASES.get(x, x) for x in extract_acronyms(b)}
    if acr_a and acr_b and (acr_a & acr_b):
        return 1.0
    base = max(similarity(a1, b1), exact_or_contains(a1, b1))
    ta = set(a1.split())
    tb = set(b1.split())
    overlap = len(ta & tb) / max(1, len(ta | tb))
    distinctive_overlap = len({t for t in ta if t not in GENERIC_TOKENS} & {t for t in tb if t not in GENERIC_TOKENS})
    if distinctive_overlap == 0 and base < 0.9:
        base *= 0.65
    return max(base, overlap)

def address_similarity(a: str, b: str) -> float:
    a1 = normalize_text(a)
    b1 = normalize_text(b)
    if not a1 or not b1:
        return 0.0
    base = max(similarity(a1, b1), exact_or_contains(a1, b1))
    ta = token_set(a1)
    tb = token_set(b1)
    overlap = len(ta & tb) / max(1, len(ta | tb))
    numeric_a = set(re.findall(r'\b\d+(?:[-/]\d+)?\b', a1))
    numeric_b = set(re.findall(r'\b\d+(?:[-/]\d+)?\b', b1))

    def normalize_num_set(nums):
        return {str(int(n.split('-')[0])) for n in nums if n.isdigit()}
    norm_a = normalize_num_set(numeric_a)
    norm_b = normalize_num_set(numeric_b)
    if (norm_a and not norm_b) or (norm_b and not norm_a):
        base *= ADDR_SIM["no_numeric_penalty"]
    elif norm_a and norm_b and not (norm_a & norm_b):
        base *= ADDR_SIM["numeric_mismatch_penalty"]
    first_num_a = re.search(r'\b\d+(?:[-/]\d+)?\b', a1)
    first_num_b = re.search(r'\b\d+(?:[-/]\d+)?\b', b1)
    if first_num_a and first_num_b and first_num_a.group(0) != first_num_b.group(0):
        base *= ADDR_SIM["first_num_mismatch_penalty"]
    numeric_score = 1.0 if norm_a and norm_a == norm_b else (0.5 if norm_a & norm_b else 0.0)
    return max(base * ADDR_SIM["base_weight"] + overlap * ADDR_SIM["overlap_weight"] + numeric_score * ADDR_SIM["numeric_score_weight"], base)

def facility_type_similarity(input_type: str, master_type: str) -> float:
    left = normalize_text(input_type)
    right = normalize_text(master_type)
    if not left or not right:
        return 0.0
    if left == right:
        return 1.0

    def collect_categories(text: str) -> set:
        cats = set()
        for canonical, values in FACILITY_TYPE_GROUPS.items():
            if text == canonical or text in values:
                cats.add(canonical)
            for value in values:
                if value in text:
                    cats.add(canonical)
        return cats

    left_cats = collect_categories(left)
    right_cats = collect_categories(right)
    if left_cats and right_cats:
        if left_cats & right_cats:
            return 1.0
        if 'factory' in left_cats and 'fabrication' in right_cats:
            return 0.45
        if 'fabrication' in left_cats and 'factory' in right_cats:
            return 0.45
        if 'ic assembly' in left_cats and 'factory' in right_cats:
            return 0.2
        if {'fabrication'} & left_cats and ({'office', 'headquarters', 'design'} & right_cats):
            return 0.05
        if {'factory'} & left_cats and ({'office', 'headquarters', 'design'} & right_cats):
            return 0.10
        if {'ic assembly'} & left_cats and ({'office', 'headquarters'} & right_cats):
            return 0.10
    return max(similarity(left, right), exact_or_contains(left, right))

def infer_confidence(score: float) -> str:
    for threshold, label in CONFIDENCE_BUCKETS:
        if score >= threshold:
            return label
    return 'No Match'

def detect_columns(df: pd.DataFrame, needed: List[str]) -> Dict[str, Optional[str]]:
    normalized = {clean_header(c): c for c in df.columns}
    resolved = {}
    for key in needed:
        result = None
        aliases = COLUMN_ALIASES.get(key, [key])
        for alias in aliases:
            alias_clean = clean_header(alias)
            if alias_clean in normalized:
                result = normalized[alias_clean]
                break
        if result is None:
            for alias in aliases:
                alias_clean = clean_header(alias)
                for nc, original in normalized.items():
                    if alias_clean in nc or nc in alias_clean:
                        result = original
                        logger.warning(f"Column '{original}' matched alias '{alias}' (substring)")
                        break
                if result:
                    break
        resolved[key] = result
    return resolved

def build_master_search_blob(row: pd.Series, cols: Dict[str, str]) -> str:
    parts = [
        safe_str(row.get(cols.get('company_name', ''))),
        safe_str(row.get(cols.get('group_name', ''))),
        safe_str(row.get(cols.get('facility_name', ''))),
        safe_str(row.get(cols.get('facility_type', ''))),
        safe_str(row.get(cols.get('full_address', ''))),
        safe_str(row.get(cols.get('city_name', ''))),
        safe_str(row.get(cols.get('state_name', ''))),
        safe_str(row.get(cols.get('country_name', ''))),
    ]
    return ' | '.join([p for p in parts if p])

# ------------------------------
# Feature parsing with space splitting and multiple address detection
# ------------------------------
def _is_address_like(part: str) -> bool:
    norm_part = normalize_text(part)
    if not norm_part:
        return False
    if len(norm_part.split()) <= 2 and not re.search(r'\d', norm_part):
        return False
    if re.search(r'^(no\.?|p\.?o\.?\s*box|\d)', part.strip(), re.IGNORECASE):
        return True
    if re.search(r'(road|street|avenue|boulevard|district|province|city|county|industrial|park|zone|taiwan|china|korea|japan|thailand|israel|malaysia|united states)', norm_part):
        return True
    if ',' in part and len(norm_part.split()) >= 4:
        return True
    return False

def _is_site_code_like(part: str) -> bool:
    norm_part = normalize_text(part)
    return bool(re.search(r'\b(fab\s*\d+[a-z]?|fab\d+[a-z]?|utl\s*\d+|utl\d+|t\s*\d+[a-z]?|t\d+[a-z]?|a\s*\d+[a-z]?|a\d+[a-z]?|asek)\b', norm_part))

def _address_richness(part: str) -> int:
    norm_part = normalize_text(part)
    if not norm_part:
        return 0
    score = 0
    score += len(re.findall(r'\d+', norm_part)) * 3
    score += len(re.findall(r'(road|street|avenue|boulevard|district|province|city|county|industrial|park|zone|sec|section|route|box)', norm_part)) * 4
    score += len(norm_part.split())
    return score

def _extract_entity_prefix(part: str) -> str:
    text = safe_str(part)
    if not text:
        return ''
    patterns = [
        r'(?=\bno\.?\s*\d)',
        r'(?=\bp\.?o\.?\s*box\b)',
        r'(?=\b\d{1,5}(?:[-/]\d+)?\b)',
        r'(?=,\s*\d{3,6}\b)',
    ]
    split_at = len(text)
    for pattern in patterns:
        match = re.search(pattern, text, re.IGNORECASE)
        if match:
            split_at = min(split_at, match.start())
    prefix = text[:split_at].strip(' ,|-')
    prefix_norm = normalize_text(prefix)
    if len(prefix_norm.split()) >= 2:
        return prefix
    return ''

def parse_feature_value(feature_value: str) -> Dict[str, object]:
    raw = safe_str(feature_value)
    raw = re.sub(r'(?<=[A-Za-z])(?=\d{1,4}(?:\b|[, -]))', ' ', raw)
    norm = normalize_text(raw)
    
    # Step 1: split on |, \n, ; and commas (keep original delimiters)
    parts = [p.strip() for p in re.split(r'[|\n;,]+', raw) if safe_str(p)]
    
    # Step 2: expand each part by splitting on spaces, but only if not address-like
    expanded_parts = []
    for part in parts:
        if _is_address_like(part):
            expanded_parts.append(part)
        else:
            subparts = [sub.strip() for sub in part.split() if sub.strip()]
            if len(subparts) > 1:
                expanded_parts.extend(subparts)
            else:
                expanded_parts.append(part)
    
    parts = expanded_parts

    # Identify all address-like parts (for multiple addresses)
    all_address_parts = [p for p in parts if _is_address_like(p)]
    # Identify company hints (non-address, length between 2 and 8 words, not too short)
    potential_companies = [p for p in parts if not _is_address_like(p) and 2 <= len(p.split()) <= 8 and len(p) >= 4]
    
    facility_hints = []
    address_hints = []
    entity_hints = []
    derived_parts = []
    for part in parts:
        norm_part = normalize_text(part)
        if norm_part and part:
            entity_hints.append(part)
            prefix = _extract_entity_prefix(part)
            if prefix and normalize_text(prefix) != norm_part:
                derived_parts.append(prefix)
        if _is_site_code_like(part):
            facility_hints.insert(0, part)
        if re.search(r'(factory|fabrication|assembly|site|plant|office|fab|campus|tower|amkor|ase|jcet|tsmc|globalfoundries)', norm_part):
            facility_hints.append(part)
        if _is_address_like(part):
            address_hints.append(part)

    all_entity_parts = parts + [p for p in derived_parts if normalize_text(p) not in {normalize_text(x) for x in parts}]
    
    # If multiple address parts exist, treat each separately later in mapping
    multiple_addresses = all_address_parts if len(all_address_parts) > 1 else []
    
    if address_hints:
        address = max(address_hints, key=_address_richness)
    else:
        address = max(parts, key=_address_richness) if parts else raw

    non_address_parts = [p for p in all_entity_parts if normalize_text(p) != normalize_text(address)]
    if facility_hints:
        facility = facility_hints[0]
    elif non_address_parts:
        facility = max(non_address_parts, key=lambda p: (1 if _is_site_code_like(p) else 0, len(normalize_text(p).split()), len(p)))
    else:
        facility = ''

    company_hints = []
    for part in non_address_parts[:4]:
        norm_part = normalize_text(part)
        if not norm_part:
            continue
        if len(norm_part.split()) <= 8 and len(norm_part) >= 4:
            company_hints.append(part)
        elif re.search(r'(tower|ase|amkor|jcet|texas instruments|hana|microchip|fujitsu|walsin|bizlink|utac|tsmc|globalfoundries|panasonic|stackpole|ttm|koa|onsemi)', norm_part):
            company_hints.append(part)
    if not company_hints and non_address_parts:
        first = non_address_parts[0]
        if not _is_address_like(first):
            company_hints.append(first)
    entity_hint = ''
    for part in non_address_parts + entity_hints:
        if part != address and not _is_address_like(part):
            entity_hint = part
            break
    if not entity_hint and non_address_parts:
        entity_hint = non_address_parts[0]

    country = ''
    for canonical, aliases in COUNTRY_SYNONYMS.items():
        if any(normalize_text(alias) in norm for alias in aliases):
            country = canonical
            break

    return {
        'raw': raw,
        'normalized': norm,
        'parts': parts,
        'non_address_parts': non_address_parts,
        'entity_hint': entity_hint,
        'facility_hint': facility,
        'address_hint': address,
        'company_hints': company_hints,
        'country_hint': country,
        'multiple_addresses': multiple_addresses,
        'potential_companies': potential_companies
    }

def normalize_country(value: str) -> str:
    norm = normalize_text(value)
    if not norm:
        return ''
    for canonical, aliases in COUNTRY_SYNONYMS.items():
        if norm == canonical or norm in {normalize_text(a) for a in aliases}:
            return canonical
    return norm

# ------------------------------
# Enhanced matching: company-first, multiple address support, facility mismatch comment
# ------------------------------
def score_candidate(input_ctx: Dict[str, str], cand: pd.Series, cols: Dict[str, str], input_facility_type: str = '') -> Dict:
    company = safe_str(cand.get(cols['company_name'])) if cols.get('company_name') else ''
    group_ = safe_str(cand.get(cols['group_name'])) if cols.get('group_name') else ''
    facility = safe_str(cand.get(cols['facility_name'])) if cols.get('facility_name') else ''
    fac_type = safe_str(cand.get(cols['facility_type'])) if cols.get('facility_type') else ''
    address = safe_str(cand.get(cols['full_address'])) if cols.get('full_address') else ''
    city = safe_str(cand.get(cols['city_name'])) if cols.get('city_name') else ''
    state = safe_str(cand.get(cols['state_name'])) if cols.get('state_name') else ''
    country = safe_str(cand.get(cols['country_name'])) if cols.get('country_name') else ''
    display_name = safe_str(cand.get('DisplayName', ''))

    manufacturer = input_ctx.get('manufacturer', '')
    supplier = input_ctx.get('contacted_supplier', '')
    feature_value = input_ctx.get('feature_value', '')
    facility_type_input = input_ctx.get('facility_type', '')
    parsed_feature = parse_feature_value(feature_value)

    # Company-first logic: if we have a company hint, boost company similarity
    has_company_hint = len(parsed_feature['company_hints']) > 0 or len(parsed_feature['potential_companies']) > 0

    explicit_company_inputs = [v for v in parsed_feature['company_hints'] if safe_str(v)]
    if not explicit_company_inputs and parsed_feature['potential_companies']:
        explicit_company_inputs = parsed_feature['potential_companies'][:3]
    supporting_company_inputs = [v for v in [manufacturer, supplier] if safe_str(v)]
    entity_hint = parsed_feature.get('entity_hint', '')
    entity_company_score = max(
        [company_similarity(entity_hint, company), company_similarity(entity_hint, group_)] if safe_str(entity_hint) else [0.0]
    )

    explicit_company_score = max(
        [company_similarity(v, company) for v in explicit_company_inputs] +
        [company_similarity(v, group_) for v in explicit_company_inputs] +
        [0.0]
    )
    supporting_company_score = max(
        [company_similarity(v, company) for v in supporting_company_inputs] +
        [company_similarity(v, group_) for v in supporting_company_inputs] +
        [0.0]
    )
    if explicit_company_inputs:
        company_score = max(entity_company_score, explicit_company_score, supporting_company_score * 0.25)
    else:
        company_score = supporting_company_score

    # If company hint exists, require a minimum company score for high confidence
    if has_company_hint and company_score < 0.5:
        company_score *= 0.7  # penalise weak company match

    explicit_group_score = max([company_similarity(v, group_) for v in explicit_company_inputs] + [0.0])
    supplier_group_score = max([company_similarity(v, group_) for v in [supplier] if safe_str(v)] + [0.0])
    group_score = max(explicit_group_score, supplier_group_score * (0.35 if explicit_company_inputs else 0.65))

    facility_inputs = [parsed_feature['facility_hint']] + parsed_feature.get('non_address_parts', [])[:4]
    facility_inputs = [v for v in facility_inputs if safe_str(v)]
    facility_score = max([similarity(v, facility) for v in facility_inputs] + [similarity(v, display_name) for v in facility_inputs] + [0.0])
    short_site_tokens = [
        normalize_text(v) for v in parsed_feature.get('non_address_parts', [])
        if 1 < len(normalize_text(v)) <= 5 and not re.search(r'\d', normalize_text(v))
    ]
    if short_site_tokens:
        site_blob = ' '.join([facility, display_name])
        for token in short_site_tokens:
            if exact_or_contains(token, site_blob) >= 0.92:
                facility_score = max(facility_score, 0.95)

    address_inputs = [parsed_feature['address_hint']] + parsed_feature['multiple_addresses'] + [feature_value]
    address_inputs = [v for v in address_inputs if safe_str(v)]
    full_geo = ' '.join([address, city, state, country]).strip()
    address_score = max([address_similarity(v, full_geo) for v in address_inputs] + [0.0])

    city_inputs = [input_ctx.get('city', ''), feature_value]
    state_inputs = [input_ctx.get('state', ''), feature_value]
    country_inputs = [input_ctx.get('country', ''), parsed_feature['country_hint'], feature_value]
    city_score = max([max(similarity(v, city), exact_or_contains(v, city)) for v in city_inputs if safe_str(v)] + [0.0])
    state_score = max([max(similarity(v, state), exact_or_contains(v, state)) for v in state_inputs if safe_str(v)] + [0.0])
    country_score = max([max(similarity(v, country), exact_or_contains(v, country)) for v in country_inputs if safe_str(v)] + [0.0])

    facility_type_score = facility_type_similarity(facility_type_input, fac_type)
    if not facility_type_score and fac_type and any(tok in parsed_feature['normalized'] for tok in normalize_text(fac_type).split()):
        facility_type_score = 0.6

    input_codes = extract_code_tokens(' '.join(parsed_feature['parts'][:3] + [feature_value]))
    candidate_codes = extract_code_tokens(' '.join([facility, display_name, company, group_]))
    code_match_score = 1.0 if input_codes and candidate_codes and input_codes & candidate_codes else 0.0
    company_present = company_score >= 0.72 or group_score >= 0.72
    address_present = address_score >= 0.75
    geo_present = sum(x >= 0.8 for x in [city_score, state_score, country_score])
    has_entity_signal = bool(parsed_feature.get('non_address_parts'))

    score = (
        company_score * MASTER_WEIGHTS['company_name'] +
        group_score * MASTER_WEIGHTS['group_name'] +
        facility_score * MASTER_WEIGHTS['facility_name'] +
        facility_type_score * MASTER_WEIGHTS['facility_type'] +
        address_score * MASTER_WEIGHTS['full_address'] +
        city_score * MASTER_WEIGHTS['city_name'] +
        state_score * MASTER_WEIGHTS['state_name'] +
        country_score * MASTER_WEIGHTS['country_name']
    )
    score += code_match_score * 20
    if code_match_score:
        facility_score = max(facility_score, 0.95)
    if address_score >= 0.85 and facility_type_score >= 0.8 and (code_match_score or facility_score >= 0.72):
        score += 10
    if not parsed_feature.get('non_address_parts') and address_score >= 0.82 and geo_present >= 2:
        score += 22
        score += min(3.0, facility.count(',') * 0.5 + fac_type.count(',') * 0.5 + (1.5 if len(re.findall(r'\b\d+\b', facility)) > 1 else 0.0))
    if not input_codes and address_score >= 0.85 and (company_score >= 0.7 or group_score >= 0.7):
        score += min(4.0, facility.count(',') * 1.5 + fac_type.count(',') * 0.5 + (1.5 if len(re.findall(r'\b\d+\b', facility)) > 1 else 0.0))

    penalties = 0.0
    reasons = []

    if facility_type_input and fac_type and facility_type_score < 0.35:
        penalties += PENALTIES["facility_type_conflict"]
        reasons.append(f'facility type conflict (input: {facility_type_input}, master: {fac_type})')

    if (manufacturer or supplier) and not company_present and not address_present:
        penalties += PENALTIES["company_mismatch_weak_address"]
        reasons.append('company mismatch with weak address support')

    if explicit_company_inputs and entity_company_score < 0.35 and explicit_company_score < 0.35 and address_score < 0.8:
        penalties += PENALTIES["feature_entity_no_support"]
        reasons.append('feature entity does not support candidate company')
    if explicit_company_inputs and company_score < 0.35 and group_score < 0.35 and not code_match_score and address_score < 0.9:
        penalties += PENALTIES["explicit_company_mismatch"]
        reasons.append('explicit company/entity mismatch')

    if has_entity_signal and not company_present and address_present and geo_present >= 2:
        penalties += PENALTIES["address_stronger_than_entity"]
        reasons.append('address stronger than entity signals')

    if has_entity_signal and company_present and address_score < 0.45 and facility_score < 0.45 and geo_present == 0:
        penalties += PENALTIES["company_match_no_site_evidence"]
        reasons.append('company match without site evidence')

    if input_ctx.get('country') and country and normalize_country(input_ctx['country']) != normalize_country(country):
        penalties += PENALTIES["country_conflict"]
        reasons.append('country conflict')

    if input_ctx.get('state') and state and similarity(input_ctx['state'], state) < 0.35 and city_score < 0.6:
        penalties += PENALTIES["state_conflict"]
        reasons.append('state conflict')

    if city and state and address_score >= 0.7 and (city_score < 0.4 or country_score < 0.6):
        penalties += PENALTIES["address_geography_conflict"]
        reasons.append('address and geography conflict')

    if input_codes and candidate_codes and not (input_codes & candidate_codes):
        penalties += PENALTIES["code_mismatch"]
        reasons.append('facility/site code mismatch')
    elif input_codes and not candidate_codes:
        p = PENALTIES["candidate_missing_code"] if company_present and address_score >= 0.75 else PENALTIES["candidate_missing_code_weak"]
        penalties += p
        reasons.append('candidate missing facility/site code')
    elif input_codes and code_match_score and address_score >= 0.8 and facility_type_score >= 0.8:
        penalties = max(0.0, penalties + PENALTIES["code_and_address_aligned_bonus"])
        reasons.append('site code and address aligned')

    final_score = max(0.0, min(100.0, score - penalties))
    
    # Facility mismatch comment
    facility_comment = ''
    if facility_type_input and fac_type and facility_type_score < 0.6:
        facility_comment = f'The facility not Matched (Master facility: {fac_type})'
    
    return {
        'score': round(final_score, 2),
        'raw_score': round(score, 2),
        'penalties': round(penalties, 2),
        'reasons': reasons,
        'facility_comment': facility_comment,
        'breakdown': {
            'company_name': round(company_score * MASTER_WEIGHTS['company_name'], 2),
            'group_name': round(group_score * MASTER_WEIGHTS['group_name'], 2),
            'facility_name': round(facility_score * MASTER_WEIGHTS['facility_name'], 2),
            'facility_type': round(facility_type_score * MASTER_WEIGHTS['facility_type'], 2),
            'full_address': round(address_score * MASTER_WEIGHTS['full_address'], 2),
            'city_name': round(city_score * MASTER_WEIGHTS['city_name'], 2),
            'state_name': round(state_score * MASTER_WEIGHTS['state_name'], 2),
            'country_name': round(country_score * MASTER_WEIGHTS['country_name'], 2),
            'code_match': round(code_match_score * 20, 2),
        }
    }

def choose_candidates(input_row: pd.Series, master_df: pd.DataFrame, input_cols: Dict[str, str], master_cols: Dict[str, str]) -> Dict:
    feature_value = safe_str(input_row.get(input_cols.get('feature_value', ''), ''))
    facility_type_input = safe_str(input_row.get(input_cols.get('input_facility_type', ''), ''))
    manufacturer = safe_str(input_row.get(input_cols.get('manufacturer', ''), ''))
    supplier = safe_str(input_row.get(input_cols.get('contacted_supplier', ''), ''))
    input_country = safe_str(input_row.get(input_cols.get('input_country', ''), ''))
    input_state = safe_str(input_row.get(input_cols.get('input_state', ''), ''))
    input_city = safe_str(input_row.get(input_cols.get('input_city', ''), ''))

    if not feature_value:
        return {
            'matched_ids': '', 'score': 0.0, 'confidence': 'No Match', 'needs_review': 'Yes',
            'selected_rows': [], 'match_basis': 'Blank FeatureValue', 'alternatives': '',
            'facility_comment': ''
        }

    parsed_feature = parse_feature_value(feature_value)
    country_hint = normalize_country(input_country or parsed_feature['country_hint'])
    input_ctx = {
        'feature_value': feature_value,
        'facility_type': facility_type_input,
        'manufacturer': manufacturer,
        'contacted_supplier': supplier,
        'country': country_hint,
        'state': input_state,
        'city': input_city,
    }

    master = master_df
    if country_hint:
        country_filtered = master[master['_country_norm'] == country_hint]
        if not country_filtered.empty:
            master = country_filtered

    feature_tokens = informative_tokens(feature_value)
    facility_tokens = informative_tokens(parsed_feature['facility_hint'])
    address_tokens = informative_tokens(parsed_feature['address_hint'])
    manufacturer_tokens = informative_tokens(manufacturer)
    supplier_tokens = informative_tokens(supplier)
    company_hint_tokens = set()
    for hint in parsed_feature['company_hints']:
        company_hint_tokens |= informative_tokens(hint)
    entity_hint = parsed_feature.get('entity_hint', '')
    input_codes = extract_code_tokens(' '.join(parsed_feature['parts'][:3] + [feature_value]))

    # Company-first shortlisting: if we have company hints, boost company-related tokens
    has_company = len(parsed_feature['company_hints']) > 0 or len(parsed_feature['potential_companies']) > 0

    def shortlist_score(row):
        blob_tokens = row['_search_blob_tokens']
        facility_name = safe_str(row.get(master_cols.get('facility_name', ''), ''))
        company_name = safe_str(row.get(master_cols.get('company_name', ''), ''))
        group_name = safe_str(row.get(master_cols.get('group_name', ''), ''))
        address = safe_str(row.get(master_cols.get('full_address', ''), ''))
        score = 0.0
        # Always include token overlaps
        score += len(feature_tokens & blob_tokens) * SHORTLIST_W["feature_token"]
        score += len(address_tokens & blob_tokens) * SHORTLIST_W["address_token"]
        score += len(facility_tokens & informative_tokens(facility_name)) * SHORTLIST_W["facility_token"]
        score += len(manufacturer_tokens & (informative_tokens(company_name) | informative_tokens(group_name))) * SHORTLIST_W["manufacturer_token"]
        score += len(supplier_tokens & informative_tokens(group_name)) * SHORTLIST_W["supplier_token"]
        # Company hint tokens get higher weight if company is present
        company_weight = SHORTLIST_W["company_hint_token"] * (1.5 if has_company else 1.0)
        score += len(company_hint_tokens & (informative_tokens(company_name) | informative_tokens(group_name) | informative_tokens(facility_name))) * company_weight
        
        if entity_hint and max(company_similarity(entity_hint, company_name), company_similarity(entity_hint, group_name), similarity(entity_hint, facility_name)) >= 0.85:
            score += SHORTLIST_W["entity_match_bonus"]
        if manufacturer and max(company_similarity(manufacturer, company_name), company_similarity(manufacturer, group_name)) >= 0.8 and not entity_hint:
            score += SHORTLIST_W["manufacturer_match_bonus"]
        if supplier and max(company_similarity(supplier, group_name), company_similarity(supplier, company_name)) >= 0.8 and not entity_hint:
            score += SHORTLIST_W["supplier_match_bonus"]
        if parsed_feature['facility_hint'] and max(similarity(parsed_feature['facility_hint'], facility_name), similarity(parsed_feature['facility_hint'], safe_str(row.get('DisplayName', '')))) >= 0.82:
            score += SHORTLIST_W["facility_hint_match_bonus"]
        if parsed_feature['address_hint'] and address_similarity(parsed_feature['address_hint'], address) >= 0.82:
            score += SHORTLIST_W["address_hint_match_bonus"]
        candidate_codes = extract_code_tokens(' '.join([facility_name, safe_str(row.get('DisplayName', '')), company_name, group_name]))
        if input_codes and candidate_codes and input_codes & candidate_codes:
            score += SHORTLIST_W["code_match_bonus"]
        return score

    master = master.copy()
    master['_shortlist_overlap'] = master.apply(shortlist_score, axis=1)
    short = master.sort_values(['_shortlist_overlap'], ascending=False).head(150).copy()

    scored = []
    for _, cand in short.iterrows():
        sc = score_candidate(input_ctx, cand, master_cols, facility_type_input)
        if sc['score'] >= SCORE_THRESHOLDS["shortlist_min_score"]:
            scored.append({'candidate': cand, **sc})

    scored.sort(key=lambda x: (x['score'], x['breakdown']['full_address'], x['breakdown']['company_name']), reverse=True)
    if not scored:
        return {
            'matched_ids': '', 'score': 0.0, 'confidence': 'No Match', 'needs_review': 'Yes',
            'selected_rows': [], 'match_basis': 'No candidates found', 'alternatives': '',
            'facility_comment': ''
        }

    # Handle multiple address parts: collect matches for each address separately
    all_matched_ids = []
    all_selected_rows = []
    all_facility_comments = []
    all_match_bases = []
    
    # If multiple addresses exist, map each one
    if parsed_feature['multiple_addresses']:
        for addr in parsed_feature['multiple_addresses']:
            # Create temporary input_ctx with this address
            temp_ctx = input_ctx.copy()
            temp_ctx['feature_value'] = addr
            temp_parsed = parse_feature_value(addr)
            temp_ctx['address_hint'] = addr
            # Score candidates for this address
            addr_scored = []
            for _, cand in short.iterrows():
                sc = score_candidate(temp_ctx, cand, master_cols, facility_type_input)
                if sc['score'] >= SCORE_THRESHOLDS["shortlist_min_score"]:
                    addr_scored.append({'candidate': cand, **sc})
            if addr_scored:
                addr_scored.sort(key=lambda x: x['score'], reverse=True)
                best = addr_scored[0]
                if best['score'] >= SCORE_THRESHOLDS["low_confidence"]:
                    all_matched_ids.append(str(best['candidate'][master_cols['location_id']]))
                    all_selected_rows.append(best['candidate'])
                    if best['facility_comment']:
                        all_facility_comments.append(best['facility_comment'])
                    all_match_bases.append(f'Address part: {addr[:50]} -> score {best["score"]}')
        # Deduplicate
        all_matched_ids = list(dict.fromkeys(all_matched_ids))
        if all_matched_ids:
            matched_ids = '|'.join(all_matched_ids)
            # For score, use average? We'll use max score from any match
            max_score = max([s['score'] for s in scored[:1]] + [0])
            confidence = infer_confidence(max_score)
            needs_review = 'No' if max_score >= 75 else 'Yes'
            facility_comment = '; '.join(dict.fromkeys(all_facility_comments))
            match_basis = 'Multiple addresses mapped; ' + '; '.join(all_match_bases[:3])
            return {
                'matched_ids': matched_ids,
                'score': max_score,
                'confidence': confidence,
                'needs_review': needs_review,
                'selected_rows': all_selected_rows,
                'match_basis': match_basis,
                'alternatives': '',
                'facility_comment': facility_comment
            }
    
    # Single address or fallback
    top = scored[0]
    top_score = top['score']
    top_breakdown = top['breakdown']

    # Check if multiple master rows have the same address (or very similar) and collect them
    same_address_rows = []
    for s in scored:
        if address_similarity(
            safe_str(s['candidate'].get(master_cols.get('full_address', ''), '')),
            safe_str(top['candidate'].get(master_cols.get('full_address', ''), ''))
        ) >= 0.95:
            same_address_rows.append(s)
    multi = same_address_rows if len(same_address_rows) > 1 else [top]

    matched_ids = '|'.join(str(s['candidate'][master_cols['location_id']]) for s in multi)
    facility_comment = '; '.join(dict.fromkeys([s['facility_comment'] for s in multi if s['facility_comment']]))
    
    if top_score < SCORE_THRESHOLDS["low_confidence"]:
        return {
            'matched_ids': '',
            'score': top_score,
            'confidence': infer_confidence(top_score),
            'needs_review': 'Yes',
            'selected_rows': [],
            'match_basis': 'Weak evidence below threshold',
            'alternatives': format_alternatives(scored[:3], master_cols),
            'facility_comment': facility_comment
        }

    close_alternative = len(scored) > 1 and abs(scored[1]['score'] - top_score) <= 5
    conflicting = any(reason in top['reasons'] for reason in ['company mismatch with weak address support', 'country conflict', 'state conflict'])
    needs_review = 'Yes' if top_score < SCORE_THRESHOLDS["needs_review_threshold"] or len(multi) > 1 or close_alternative or conflicting else 'No'

    basis_parts = []
    if top_breakdown['company_name'] + top_breakdown['group_name'] >= 24:
        basis_parts.append('strong company/group match')
    elif top_breakdown['company_name'] + top_breakdown['group_name'] >= 14:
        basis_parts.append('partial company/group match')

    if top_breakdown['facility_name'] >= 15:
        basis_parts.append('strong facility name match')
    elif top_breakdown['facility_name'] >= 9:
        basis_parts.append('partial facility name match')

    if top_breakdown['full_address'] >= 15:
        basis_parts.append('strong address match')
    elif top_breakdown['full_address'] >= 9:
        basis_parts.append('partial address match')

    geo_points = sum(top_breakdown[k] >= 4 for k in ['city_name', 'state_name', 'country_name'])
    if geo_points >= 2:
        basis_parts.append('geography consistent')

    if top_breakdown['facility_type'] >= 8:
        basis_parts.append('facility type aligned')
    elif facility_type_input:
        basis_parts.append('facility type weak/partial')

    if len(multi) > 1:
        basis_parts.append('multiple valid IDs for same physical site')
    if close_alternative:
        basis_parts.append('close alternative candidates exist')
    if top['reasons']:
        basis_parts.append('penalties: ' + ', '.join(top['reasons']))

    return {
        'matched_ids': matched_ids,
        'score': top_score,
        'confidence': infer_confidence(top_score),
        'needs_review': needs_review,
        'selected_rows': [s['candidate'] for s in multi],
        'match_basis': '; '.join(basis_parts) if basis_parts else 'best overall weighted match',
        'alternatives': format_alternatives(scored[1:4], master_cols),
        'facility_comment': facility_comment
    }

def format_alternatives(scored_list: List[Dict], master_cols: Dict[str, str]) -> str:
    items = []
    for s in scored_list:
        cand = s['candidate']
        loc = safe_str(cand.get(master_cols.get('location_id', '')))
        comp = safe_str(cand.get(master_cols.get('company_name', '')))
        fac = safe_str(cand.get(master_cols.get('facility_name', '')))
        city = safe_str(cand.get(master_cols.get('city_name', '')))
        if loc:
            items.append(f"{loc} ({comp} | {fac} | {city} | score={s['score']})")
    return ' || '.join(items)

# ------------------------------
# Abbreviation-based direct mapping (fuzzy + contains)
# ------------------------------
def load_abbreviation_map(abbr_path: Optional[Path]) -> Tuple[Optional[Dict], Optional[List[Tuple[str, Dict]]]]:
    if not abbr_path or not abbr_path.exists():
        return None, None
    try:
        if abbr_path.suffix.lower() == '.csv':
            df = pd.read_csv(abbr_path)
        else:
            df = pd.read_excel(abbr_path)
        df.columns = [clean_header(c) for c in df.columns]
        expected = ['locationabbreviation', 'sitelevel', 'companylocationid', 'country', 'state', 'city', 'manufacturingcompany']
        for exp in expected:
            if exp not in df.columns:
                raise ValueError(f"Abbreviation file missing required column: {exp} (normalised). Found: {list(df.columns)}")
        exact_map = {}
        fuzzy_list = []
        for _, row in df.iterrows():
            key_raw = safe_str(row['locationabbreviation'])
            if not key_raw:
                continue
            norm_key = normalize_text(key_raw)
            mapping = {
                'SiteLevel': row['sitelevel'],
                'CompanyLocationID': row['companylocationid'],
                'Country': row['country'],
                'State': row['state'],
                'City': row['city'],
                'ManufacturingCompany': row['manufacturingcompany'],
                'original_abbreviation': key_raw
            }
            exact_map[norm_key] = mapping
            fuzzy_list.append((norm_key, mapping))
        logger.info(f"Loaded {len(exact_map)} abbreviation mappings from {abbr_path}")
        return exact_map, fuzzy_list
    except Exception as e:
        logger.error(f"Failed to load abbreviation file: {e}")
        raise

def find_best_abbreviation_match(feature_value: str, exact_map: Dict[str, Dict], fuzzy_list: List[Tuple[str, Dict]], threshold: int = ABBREVIATION_MATCH_THRESHOLD) -> Optional[Dict]:
    norm_feature = normalize_text(feature_value)
    if norm_feature in exact_map:
        logger.debug(f"Exact abbreviation match: '{feature_value}' -> {exact_map[norm_feature]['CompanyLocationID']}")
        return exact_map[norm_feature]

    if fuzzy_list and process is not None:
        keys = [k for k, _ in fuzzy_list]
        result = process.extractOne(norm_feature, keys, scorer=fuzz.ratio, score_cutoff=threshold)
        if result:
            best_key, score, _ = result
            for k, mapping in fuzzy_list:
                if k == best_key:
                    logger.info(f"Fuzzy abbreviation match: '{feature_value}' -> '{mapping['original_abbreviation']}' (score={score})")
                    return mapping

    for norm_key, mapping in fuzzy_list:
        if norm_feature in norm_key or norm_key in norm_feature:
            logger.info(f"Contains abbreviation match: '{feature_value}' contains/is contained in '{mapping['original_abbreviation']}'")
            return mapping

    return None

# ------------------------------
# Main pipeline with abbreviation and formatting
# ------------------------------
def map_locations(master_path: Path, input_path: Path, output_path: Path,
                  master_sheet: Optional[str] = None, input_sheet: Optional[str] = None,
                  abbreviation_path: Optional[Path] = None,
                  progress_callback=None):
    try:
        master_sheet_name = 0 if master_sheet in (None, '') else master_sheet
        input_sheet_name = 0 if input_sheet in (None, '') else input_sheet
        if master_path.suffix.lower() == '.csv':
            master_df = pd.read_csv(master_path)
        else:
            master_df = pd.read_excel(master_path, sheet_name=master_sheet_name)

        if input_path.suffix.lower() == '.csv':
            input_df = pd.read_csv(input_path)
        else:
            input_df = pd.read_excel(input_path, sheet_name=input_sheet_name)

        exact_abbr_map, fuzzy_abbr_list = load_abbreviation_map(abbreviation_path) if abbreviation_path else (None, None)

        logger.info(f"Loaded master: {len(master_df)} rows, input: {len(input_df)} rows")
        if exact_abbr_map:
            logger.info(f"Abbreviation mappings available: {len(exact_abbr_map)}")
    except Exception as e:
        logger.error(f"Failed to load files: {e}")
        raise

    master_cols = detect_columns(master_df, [
        'location_id', 'company_name', 'group_name', 'facility_name', 'facility_type',
        'country_name', 'full_address', 'state_name', 'city_name'
    ])
    input_cols = detect_columns(input_df, [
        'input_row_id', 'input_facility_type', 'feature_value',
        'manufacturer', 'contacted_supplier', 'input_country', 'input_state', 'input_city'
    ])

    required_master = ['location_id', 'company_name', 'group_name', 'facility_name', 'facility_type', 'country_name', 'full_address', 'state_name', 'city_name']
    missing_master = [k for k in required_master if not master_cols.get(k)]
    missing_input = [k for k in ['feature_value', 'input_facility_type'] if not input_cols.get(k)]

    if missing_master:
        raise ValueError(f'Missing required master columns: {missing_master}')
    if missing_input:
        raise ValueError(f'Missing required input columns: {missing_input}')

    master_df['_search_blob_norm'] = master_df.apply(lambda r: normalize_text(build_master_search_blob(r, master_cols)), axis=1)
    master_df['_search_blob_tokens'] = master_df['_search_blob_norm'].apply(informative_tokens)
    master_df['_country_norm'] = master_df[master_cols['country_name']].apply(normalize_country)

    results = []
    decision_cache = {}
    total_rows = len(input_df)

    for idx, row in input_df.iterrows():
        if progress_callback:
            progress_callback(idx + 1, total_rows)

        feature_value = safe_str(row.get(input_cols.get('feature_value', ''), ''))
        detected_row_id = safe_str(row.get(input_cols.get('input_row_id'))) if input_cols.get('input_row_id') else ''
        row_id = detected_row_id if detected_row_id else str(idx + 1)

        # Try abbreviation match first (exact, fuzzy, contains)
        abbr_match = None
        if exact_abbr_map is not None:
            abbr_match = find_best_abbreviation_match(feature_value, exact_abbr_map, fuzzy_abbr_list)

        if abbr_match is not None:
            result_row = dict(row)
            result_row.update({
                'Input_Row_ID': row_id,
                'Location Abbreviation': feature_value,
                'FacilityType': safe_str(row.get(input_cols.get('input_facility_type', ''), '')),
                'FeatureValue': feature_value,
                'Matched_Location_ID': abbr_match['CompanyLocationID'],
                'Match_Confidence_Score': 100.0,
                'Confidence_Level': 'High confidence (abbreviation match)',
                'Matched_Company_Name': abbr_match['ManufacturingCompany'],
                'Matched_GroupName': '',
                'Matched_Facility_Name': '',
                'Matched_Facility_Type': abbr_match['SiteLevel'],
                'Matched_Full_Address': '',
                'Matched_City_name': abbr_match['City'],
                'Matched_State_Name': abbr_match['State'],
                'Matched_CountryName': abbr_match['Country'],
                'Match_Basis': f"Abbreviation match (exact/fuzzy/contains) - original: {abbr_match.get('original_abbreviation', '')}",
                'Alternative_Candidates': '',
                'Needs_Review': 'No',
                'SiteLevel': abbr_match['SiteLevel'],
                'CompanyLocationID': abbr_match['CompanyLocationID'],
                'Country': abbr_match['Country'],
                'State': abbr_match['State'],
                'City': abbr_match['City'],
                'ManufacturingCompany': abbr_match['ManufacturingCompany'],
                'Facility_Comment': ''
            })
            results.append(result_row)
            continue

        # No abbreviation match – use master matching
        cache_key = (
            feature_value,
            safe_str(row.get(input_cols.get('input_facility_type', ''), '')),
            safe_str(row.get(input_cols.get('manufacturer', ''), '')),
            safe_str(row.get(input_cols.get('contacted_supplier', ''), '')),
            safe_str(row.get(input_cols.get('input_country', ''), '')),
            safe_str(row.get(input_cols.get('input_state', ''), '')),
            safe_str(row.get(input_cols.get('input_city', ''), '')),
        )
        if cache_key not in decision_cache:
            decision_cache[cache_key] = choose_candidates(row, master_df, input_cols, master_cols)
        decision = decision_cache[cache_key]
        selected = decision['selected_rows']

        def get_selected(col_key: str) -> str:
            if not selected or not master_cols.get(col_key):
                return ''
            values = [safe_str(sel.get(master_cols[col_key])) for sel in selected]
            values = [v for v in values if v]
            return '|'.join(dict.fromkeys(values))

        # If multiple matched IDs exist, we need to combine geography and facility type appropriately
        matched_ids = decision['matched_ids']
        if '|' in matched_ids:
            # Multiple IDs: gather info from each selected row
            companies = []
            groups = []
            facilities = []
            facility_types = []
            addresses = []
            cities = []
            states = []
            countries = []
            for sel in selected:
                companies.append(safe_str(sel.get(master_cols['company_name'])))
                groups.append(safe_str(sel.get(master_cols['group_name'])))
                facilities.append(safe_str(sel.get(master_cols['facility_name'])))
                facility_types.append(safe_str(sel.get(master_cols['facility_type'])))
                addresses.append(safe_str(sel.get(master_cols['full_address'])))
                cities.append(safe_str(sel.get(master_cols['city_name'])))
                states.append(safe_str(sel.get(master_cols['state_name'])))
                countries.append(safe_str(sel.get(master_cols['country_name'])))
            matched_company = '|'.join(dict.fromkeys([c for c in companies if c]))
            matched_group = '|'.join(dict.fromkeys([g for g in groups if g]))
            matched_facility = '|'.join(dict.fromkeys([f for f in facilities if f]))
            matched_facility_type = '|'.join(dict.fromkeys([ft for ft in facility_types if ft]))
            matched_address = '|'.join(dict.fromkeys([a for a in addresses if a]))
            matched_city = '|'.join(dict.fromkeys([c for c in cities if c]))
            matched_state = '|'.join(dict.fromkeys([s for s in states if s]))
            matched_country = '|'.join(dict.fromkeys([co for co in countries if co]))
        else:
            matched_company = get_selected('company_name')
            matched_group = get_selected('group_name')
            matched_facility = get_selected('facility_name')
            matched_facility_type = get_selected('facility_type')
            matched_address = get_selected('full_address')
            matched_city = get_selected('city_name')
            matched_state = get_selected('state_name')
            matched_country = get_selected('country_name')

        result_row = dict(row)
        result_row.update({
            'Input_Row_ID': row_id,
            'Location Abbreviation': feature_value,
            'FacilityType': safe_str(row.get(input_cols.get('input_facility_type', ''), '')),
            'FeatureValue': feature_value,
            'Matched_Location_ID': decision['matched_ids'],
            'Match_Confidence_Score': decision['score'],
            'Confidence_Level': decision['confidence'],
            'Matched_Company_Name': matched_company,
            'Matched_GroupName': matched_group,
            'Matched_Facility_Name': matched_facility,
            'Matched_Facility_Type': matched_facility_type,
            'Matched_Full_Address': matched_address,
            'Matched_City_name': matched_city,
            'Matched_State_Name': matched_state,
            'Matched_CountryName': matched_country,
            'Match_Basis': decision['match_basis'],
            'Alternative_Candidates': decision['alternatives'],
            'Needs_Review': decision['needs_review'],
            'SiteLevel': matched_facility_type,
            'CompanyLocationID': decision['matched_ids'],
            'Country': matched_country,
            'State': matched_state,
            'City': matched_city,
            'ManufacturingCompany': matched_company or matched_group,
            'Facility_Comment': decision.get('facility_comment', '')
        })
        results.append(result_row)

    output_df = pd.DataFrame(results)

    updated_mapping_cols = [
        'Input_Row_ID', 'Location Abbreviation', 'FacilityType', 'SiteLevel', 'CompanyLocationID',
        'Country', 'State', 'StateProvince', 'City', 'ManufacturingCompany',
        'Match_Confidence_Score', 'Confidence_Level', 'Needs_Review', 'Facility_Comment'
    ]
    audit_cols = [
        'Input_Row_ID', 'FacilityType', 'FeatureValue', 'Matched_Location_ID', 'Match_Confidence_Score',
        'Confidence_Level', 'Matched_Company_Name', 'Matched_GroupName', 'Matched_Facility_Name',
        'Matched_Facility_Type', 'Matched_Full_Address', 'Matched_City_name', 'Matched_State_Name',
        'Matched_CountryName', 'Match_Basis', 'Alternative_Candidates', 'Needs_Review', 'Facility_Comment'
    ]
    updated_mapping_df = output_df.reindex(columns=updated_mapping_cols)
    audit_df = output_df.reindex(columns=audit_cols)

    summary = pd.DataFrame([
        {'Metric': 'Input rows', 'Value': len(output_df), 'Unnamed: 2': '', 'Unnamed: 3': ''},
        {'Metric': 'Matched rows (any)', 'Value': int((output_df['Matched_Location_ID'].fillna('') != '').sum()), 'Unnamed: 2': '', 'Unnamed: 3': ''},
        {'Metric': 'No Match rows', 'Value': int((output_df['Matched_Location_ID'].fillna('') == '').sum()), 'Unnamed: 2': '', 'Unnamed: 3': ''},
        {'Metric': 'Needs Review rows', 'Value': int((output_df['Needs_Review'] == 'Yes').sum()), 'Unnamed: 2': '', 'Unnamed: 3': ''},
    ])

    # Write Excel with conditional formatting
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        updated_mapping_df.to_excel(writer, index=False, sheet_name='Updated_Mapping')
        audit_df.to_excel(writer, index=False, sheet_name='Audit_Details')
        summary.to_excel(writer, index=False, sheet_name='Summary')
        
        # Apply conditional formatting to Confidence_Level column in Updated_Mapping sheet
        workbook = writer.book
        sheet = writer.sheets['Updated_Mapping']
        # Find column index for Confidence_Level (case insensitive)
        col_idx = None
        for cell in sheet[1]:  # header row
            if cell.value and 'confidence_level' in str(cell.value).lower():
                col_idx = cell.column
                break
        if col_idx:
            # Define fills
            green_fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")   # light green
            yellow_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")  # gold/yellow
            orange_fill = PatternFill(start_color="FF8C00", end_color="FF8C00", fill_type="solid")  # dark orange
            red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")     # red
            for row in range(2, sheet.max_row + 1):
                cell = sheet.cell(row, col_idx)
                val = str(cell.value).lower() if cell.value else ''
                if 'high' in val:
                    cell.fill = green_fill
                elif 'medium' in val:
                    cell.fill = yellow_fill
                elif 'low' in val or 'review' in val:
                    cell.fill = orange_fill
                elif 'no match' in val:
                    cell.fill = red_fill

    logger.info(f'Wrote output: {output_path}')

# ------------------------------
# CLI and GUI (unchanged)
# ------------------------------
if __name__ == '__main__':
    parser = argparse.ArgumentParser(description='Map input locations to master Location_IDs.')
    parser.add_argument('--master', required=False, help='Path to master file')
    parser.add_argument('--input', required=False, help='Path to input mapping file')
    parser.add_argument('--output', required=False, help='Path to output xlsx file')
    parser.add_argument('--abbreviation', required=False, help='Path to abbreviation mapping file (optional)')
    parser.add_argument('--master-sheet', default=None, help='Optional master sheet name')
    parser.add_argument('--input-sheet', default=None, help='Optional input sheet name')
    args = parser.parse_args()

    def run_cli():
        if not args.master or not args.input or not args.output:
            raise SystemExit('CLI mode requires --master, --input, and --output.')
        try:
            from tqdm import tqdm
            total_rows = pd.read_excel(args.input, sheet_name=args.input_sheet or 0).shape[0] if not args.input.endswith('.csv') else pd.read_csv(args.input).shape[0]
            pbar = tqdm(total=total_rows, desc="Processing rows")
            def progress_callback(current, total):
                pbar.update(1)
            map_locations(
                master_path=Path(args.master),
                input_path=Path(args.input),
                output_path=Path(args.output),
                master_sheet=args.master_sheet,
                input_sheet=args.input_sheet,
                abbreviation_path=Path(args.abbreviation) if args.abbreviation else None,
                progress_callback=progress_callback
            )
            pbar.close()
        except ImportError:
            logger.warning("tqdm not installed, progress bar disabled.")
            map_locations(
                master_path=Path(args.master),
                input_path=Path(args.input),
                output_path=Path(args.output),
                master_sheet=args.master_sheet,
                input_sheet=args.input_sheet,
                abbreviation_path=Path(args.abbreviation) if args.abbreviation else None
            )

    class LocationMappingApp:
        def __init__(self) -> None:
            self.root = tk.Tk()
            self.root.title('Location Mapping App')
            self.root.geometry('760x520')
            self.root.minsize(720, 480)

            self.master_path = tk.StringVar()
            self.input_path = tk.StringVar()
            self.abbreviation_path = tk.StringVar()
            self.output_path = tk.StringVar()
            self.status = tk.StringVar(value='Select the master file, input file, and output file to start.')
            self.running = False
            self.progress_var = tk.IntVar()
            self._build_ui()

        def _build_ui(self) -> None:
            outer = ttk.Frame(self.root, padding=16)
            outer.pack(fill='both', expand=True)

            title = ttk.Label(outer, text='Location Mapping Desktop App', font=('Segoe UI', 14, 'bold'))
            title.grid(row=0, column=0, columnspan=3, sticky='w', pady=(0, 6))

            subtitle = ttk.Label(outer, text='Browse for the required files. Abbreviation file is optional but improves speed.')
            subtitle.grid(row=1, column=0, columnspan=3, sticky='w', pady=(0, 14))

            self._add_file_row(outer, 2, 'Master file', self.master_path,
                               lambda: self._browse_file(self.master_path, 'Select master file'))
            self._add_file_row(outer, 3, 'Input file', self.input_path,
                               lambda: self._browse_file(self.input_path, 'Select input mapping file'))
            self._add_file_row(outer, 4, 'Abbreviation file (optional)', self.abbreviation_path,
                               lambda: self._browse_file(self.abbreviation_path, 'Select abbreviation mapping file'))
            self._add_file_row(outer, 5, 'Output file', self.output_path, self._browse_output)

            self.run_button = ttk.Button(outer, text='Run Mapping', command=self._start_mapping)
            self.run_button.grid(row=6, column=0, sticky='w', pady=(18, 8))

            ttk.Button(outer, text='Exit', command=self.root.destroy).grid(row=6, column=1, sticky='w', pady=(18, 8))

            self.progress_bar = ttk.Progressbar(outer, orient='horizontal', length=400, mode='determinate', variable=self.progress_var)
            self.progress_bar.grid(row=7, column=0, columnspan=2, sticky='ew', pady=(10, 5))

            status_label = ttk.Label(outer, textvariable=self.status, foreground='#0f4c81', wraplength=700, justify='left')
            status_label.grid(row=8, column=0, columnspan=3, sticky='ew', pady=(12, 0))

            outer.columnconfigure(1, weight=1)

        def _add_file_row(self, parent, row_num: int, label: str, variable: tk.StringVar, browse_command):
            ttk.Label(parent, text=label).grid(row=row_num, column=0, sticky='w', pady=6)
            ttk.Entry(parent, textvariable=variable, width=60).grid(row=row_num, column=1, sticky='ew', padx=(10, 10), pady=6)
            ttk.Button(parent, text='Browse...', command=browse_command).grid(row=row_num, column=2, sticky='ew', pady=6)

        def _browse_file(self, target: tk.StringVar, title: str):
            path = filedialog.askopenfilename(title=title, filetypes=[('Excel/CSV files', '*.xlsx *.xls *.xlsm *.csv'), ('All files', '*.*')])
            if path:
                target.set(path)
                if target is self.input_path and not self.output_path.get():
                    suggested = str(Path(path).with_name(Path(path).stem + '_mapped.xlsx'))
                    self.output_path.set(suggested)

        def _browse_output(self):
            suggested_name = 'mapped_output.xlsx'
            if self.input_path.get():
                suggested_name = Path(self.input_path.get()).stem + '_mapped.xlsx'
            path = filedialog.asksaveasfilename(title='Choose output file', defaultextension='.xlsx',
                                                initialfile=suggested_name, filetypes=[('Excel files', '*.xlsx')])
            if path:
                self.output_path.set(path)

        def _set_running_state(self, running: bool):
            self.running = running
            self.run_button.config(state='disabled' if running else 'normal')

        def _start_mapping(self):
            if self.running:
                return
            master = self.master_path.get().strip()
            input_file = self.input_path.get().strip()
            output = self.output_path.get().strip()
            if not master or not input_file or not output:
                messagebox.showwarning('Missing files', 'Please choose the master file, input file, and output file.')
                return

            abbr = self.abbreviation_path.get().strip()
            if not abbr:
                abbr = None

            self._set_running_state(True)
            self.progress_var.set(0)
            self.status.set('Mapping is running...')
            worker = threading.Thread(target=self._run_mapping_worker,
                                      args=(master, input_file, output, abbr),
                                      daemon=True)
            worker.start()

        def _run_mapping_worker(self, master: str, input_file: str, output: str, abbr: Optional[str]):
            try:
                try:
                    if input_file.endswith('.csv'):
                        total = pd.read_csv(input_file).shape[0]
                    else:
                        total = pd.read_excel(input_file).shape[0]
                except:
                    total = 100
                self.root.after(0, lambda: self.progress_bar.config(maximum=total))

                def progress_callback(current, total):
                    self.root.after(0, lambda: self.progress_var.set(current))

                map_locations(
                    master_path=Path(master),
                    input_path=Path(input_file),
                    output_path=Path(output),
                    master_sheet=None,
                    input_sheet=None,
                    abbreviation_path=Path(abbr) if abbr else None,
                    progress_callback=progress_callback
                )
                self.root.after(0, lambda: self._on_success(output))
            except Exception as exc:
                error_text = ''.join(traceback.format_exception_only(type(exc), exc)).strip()
                self.root.after(0, lambda: self._on_error(error_text))

        def _on_success(self, output: str):
            self._set_running_state(False)
            self.status.set(f'Mapping completed successfully. Output saved to: {output}')
            messagebox.showinfo('Completed', f'Mapping completed successfully.\n\nOutput:\n{output}')

        def _on_error(self, error_text: str):
            self._set_running_state(False)
            self.status.set('Mapping failed. Review the error message and try again.')
            messagebox.showerror('Mapping failed', error_text)

        def run(self):
            self.root.mainloop()

    if args.master or args.input or args.output:
        run_cli()
    else:
        LocationMappingApp().run()
